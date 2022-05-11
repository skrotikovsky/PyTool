import datetime
import os
import pyexcel
import openpyxl
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font, GradientFill
)
from openpyxl.styles import Border, Side

array_of_colors = ['CCD1BF', 'D4D2AA', 'DFE2E4', 'C9D5D1', 'D5CAAF', 'CFB677', 'BED2B8', 'ACC1CB', 'CECEA9', 'A1BCCB', 'D6DCC6']
path_exel = r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/Просмотр/Коллизии/2022.4.24/3.xlsx"
# path_exel = r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/Просмотр/Коллизии/2022.4.24"
path_of_main = r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples"


def get_xlsx_in_dir(path):
    names = os.listdir(path)
    file_list = []
    for i in range(len(names)):
        if names[i].endswith(".xlsx"):
            names[i] = path + '/' + names[i]
            file_list.append(names[i])
    #    print(file_list)
    return file_list


def find_xlsx(path):
    html_dirs = []
    for root, dirs, files in os.walk(path):  # поиск всех html файлов в указанной директории получаем массив директорий
        for file in files:
            html_dirs.append(root)
    #    print(list(set(html_dirs)))
    return list(set(html_dirs))


def get_exel_array(path):  # представляет отчет в виде двумерного массива
    return pyexcel.get_array(file_name=path)


def get_otchet_rows_dict(path):  # из названия файлов достает марку и делает приписку а так же достает новые коллизии
    # и неразрешенные коллизии а так же схлопывает строки если марки одинаковые
    added_marks = []
    rows_array = []
    data = get_exel_array(path)
    for i in data[1:len(data) - 1]:
        marks = i[0].split('.')
        i[0] = marks[len(marks) - 1]
        splited_mark = i[0].split('_')
        replaced_mark = f'{splited_mark[1]}_{splited_mark[0]}'

        if replaced_mark in added_marks:
            rows_array.append([f'Кол-во конфликтов между {replaced_mark}', i[1], i[4]])
            added_marks.append(replaced_mark)
        else:
            rows_array.append([f'Кол-во конфликтов между {i[0]}', i[1], i[4]])
            added_marks.append(i[0])
    for i in range(len(rows_array)):
        for j in range(3):
            if rows_array[i][j] == '' or rows_array[i][j] == 'Коллизий не обнаружено':
                rows_array[i][j] = 0
    marks_and_rows_dict = {}
    for value in rows_array:
        if value[0] not in marks_and_rows_dict.keys():
            marks_and_rows_dict.update({value[0]: [value[1], value[2]]})
        else:
            mark_value = marks_and_rows_dict[value[0]]
            marks_and_rows_dict[value[0]] = [int(mark_value[0]) + int(value[1]), int(mark_value[1]) + int(value[2])]
    return marks_and_rows_dict


def get_otchet_marks_array(path):  # достает только марки и делает приписку
    marks_array = []
    added_marks = []
    data = get_exel_array(path)
    for i in data[1:len(data) - 1]:
        marks = i[0].split('.')
        i[0] = marks[len(marks) - 1]
        splited_mark = i[0].split('_')
        replaced_mark = f'{splited_mark[1]}_{splited_mark[0]}'
        if replaced_mark in added_marks:
            marks_array.append(f'Кол-во конфликтов между {replaced_mark}')
            added_marks.append(replaced_mark)
        else:
            marks_array.append(f'Кол-во конфликтов между {i[0]}')
            added_marks.append(i[0])
    return list(set(marks_array))


def get_main_otchet_array(main_path):  # читает главный отчет и представляет его в виде двумерного массива
    return pyexcel.get_array(file_name=get_xlsx_in_dir(main_path)[0])


def get_main_otchet_marks(main_path):  # возвращает марки из главного отчета
    return set(list(map(lambda x: x[0], get_main_otchet_array(main_path)))[2:])


def get_new_marks(path, main_path):  # возвращает новые марки
    marks = set(get_otchet_marks_array(path))
    main_marks = get_main_otchet_marks(main_path)
    difference = marks - main_marks
    return list(difference)


def get_rows_for_empty_list(path, worksheet):  # добавляет марки из отчета в главный отчет
    marks = get_otchet_marks_array(path)
    sorted_marks = sorted(marks)
    worksheet.cell(row=0 + 1, column=0 + 1).value = 'Дата'
    worksheet.cell(row=0 + 2, column=0 + 1).value = 'Конфликты'
    for index, value in enumerate(sorted_marks):
        worksheet.cell(row=index + 3, column=0 + 1).value = value


def get_main_marks_and_rows(main_path):  # превращает главный exel отчет в пары ключ(марка): значение(строка после)
    main_rows_dict = {}
    for i in get_main_otchet_array(main_path):
        main_rows_dict.update({i[0]: i[1:]})
    return main_rows_dict


def moved_right_rows(path, main_path):  # добвляет в начале строки 2 пустых элемента в которые будут вписаны данные
    # новой проверки
    # wb = openpyxl.load_workbook(r"c:/users/skrut/onedrive/рабочий стол/exelexamples/kt101r_главный отчет.xlsx")
    # worksheet = wb['openpyxl']
    # wb.save(r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/KT101R_Главный отчет.xlsx")
    # print(worksheet.cell(row=1, column=2).value)
    moved_rows = get_main_marks_and_rows(main_path)
    # len_of_line = len(moved_rows[''])

    for i in moved_rows.keys():
        moved_rows[i] = [0, 0] + moved_rows[i]
    new_marks = get_new_marks(path, main_path)
    print(new_marks)
    if new_marks:
        for i in new_marks:
            moved_rows.update({i: [0, 0]})
    return moved_rows


def write_row(row, row_num, column_start, worksheet):  # функция написания строки в лист(worksheet) в строку номер (row
    # num) а сама
    # строка = row
    for i, value in enumerate(row):
        worksheet.cell(row=row_num, column=i + column_start).value = row[i]


def paint_row(row, row_num, color, worksheet):
    for i in range(len(row) + 1):
        worksheet.cell(row=row_num, column=i + 1).fill = PatternFill('solid',
                                                                     fgColor=color)
        thin = Side(border_style="thin", color="2E3234")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        worksheet.cell(row=row_num, column=i + 1).border = border


def write_if_main_is_empty(path, main_path):  # если главный отчет пустой - заполняет его марками в первом столбце(теми
    # которые были в начальных данных в алфавитном порядке марок) и вызывает функцию заполнения
    # marks_and_rows = get_marks_and_row_dict(path, main_path)
    # sorted_marks = sorted(marks_and_rows.keys())
    wb = openpyxl.load_workbook(r"c:/users/skrut/onedrive/рабочий стол/exelexamples/kt101r_главный отчет.xlsx")
    worksheet = wb['openpyxl']
    get_rows_for_empty_list(path, worksheet)
    wb.save(r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/KT101R_Главный отчет.xlsx")
    write_if_data_exists(path, main_path)


def write_if_data_exists(path, main_path):  # если главный отчет уже заполнен - добавляет данные по новой проверке
    # в начало списка проверок
    main_array = get_main_otchet_array(main_path)
    wb = openpyxl.load_workbook(r"c:/users/skrut/onedrive/рабочий стол/exelexamples/kt101r_главный отчет.xlsx")
    worksheet = wb['openpyxl']
    current_otchet_rows = get_otchet_rows_dict(path)
    line_1 = main_array[0]
    line_2 = main_array[1]
    line_1.insert(1, '')
    line_1.insert(1, f'{datetime.date.today()}')
    line_2.insert(1, 'Конфликты')
    line_2.insert(1, 'Дата')
    moved_main_otchet_rows = moved_right_rows(path, main_path)
    sorted_otchet_marks = sorted(moved_main_otchet_rows.keys())
    keys_and_colors = {}
    counter = 0
    for index, value in enumerate(sorted_otchet_marks):
        if value == 'Дата' or value == 'Конфликты' or value == '':
            continue
        if value in current_otchet_rows.keys():
            moved_main_otchet_rows[value][0] = current_otchet_rows[value][0]
            moved_main_otchet_rows[value][1] = current_otchet_rows[value][1]
        else:
            moved_main_otchet_rows[value][0] = ''
            moved_main_otchet_rows[value][1] = ''
        if counter == len(array_of_colors):
            counter = 0
        worksheet.cell(row=index+1, column=0 + 1).value = value
        keys_and_colors.update({value: array_of_colors[counter]})
        counter += 1
        # worksheet.cell(row=index + 3, column=0 + 1).fill = PatternFill('solid', fgColor="DDDDDD")

    for value in range(len(moved_main_otchet_rows.keys())):
        row_key = worksheet.cell(row=value + 1, column=0 + 1).value
        if row_key != 'Дата' and row_key != 'Конфликты' and row_key is not None:
            write_row(moved_main_otchet_rows[row_key], value + 1, 2, worksheet)
            paint_row(moved_main_otchet_rows[row_key], value + 1, keys_and_colors[row_key], worksheet)
    write_row(line_1, 0 + 1, 1, worksheet)
    write_row(line_2, 1 + 1, 1, worksheet)
    paint_row(line_1, 0 + 1, '92FF88', worksheet)
    paint_row(line_1, 1 + 1, '92FF88', worksheet)
    wb.save(r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/KT101R_Главный отчет.xlsx")


def write_data_in_main_otchet(main_path):  # заполняет данными в зависимости от того заполнен ли главный отчет
    # или он пуст
    main_array = get_main_otchet_array(main_path)
    for i in main_array:
        if not i:
            continue
        else:
            return write_if_data_exists
    return write_if_main_is_empty


write_data_in_main_otchet(path_of_main)(path_exel, path_of_main)
