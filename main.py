from PyQt5 import QtCore
from PyQt5.QtWidgets import QFileDialog, QTextEdit, QMessageBox
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton
import sys
import datetime
import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Border, Side
)

array_of_colors = ['CCD1BF', 'D4D2AA', 'DFE2E4', 'C9D5D1', 'D5CAAF', 'CFB677', 'BED2B8', 'ACC1CB', 'CECEA9', 'A1BCCB',
                   'D6DCC6']
# main_otchet_file = r'C:\Users\skrut\OneDrive\Рабочий стол\exelExamples\KT101R_Главный отчет.xlsx'
# exel_otchet_file = r'C:\Users\skrut\OneDrive\Рабочий стол\exelExamples\Просмотр\Коллизии\2022.4.24\1.xlsx'
main_otchet_file = 'C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/KT101R_Главный отчет.xlsx'
exel_otchet_file = 'C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/Просмотр/Коллизии/2022.4.24/1.xlsx'
print(main_otchet_file)
print(exel_otchet_file)


def get_xlsx_in_dir(path):
    names = os.listdir(path)
    file_list = []
    for i in range(len(names)):
        if names[i].endswith(".xlsx"):
            names[i] = path + '/' + names[i]
            file_list.append(names[i])
    return file_list


def find_xlsx(path):
    html_dirs = []
    for root, dirs, files in os.walk(path):  # поиск всех html файлов в указанной директории получаем массив директорий
        for file in files:
            html_dirs.append(root)
    return list(set(html_dirs))


def get_exel_array():  # представляет отчет в виде двумерного массива
    wb = openpyxl.load_workbook(exel_otchet_file, data_only=True)
    worksheet = wb['Лист1']
    exel_array = []
    for i in range(worksheet.max_row):
        exel_array.append([])
        for j in range(worksheet.max_column):
            value = worksheet.cell(row=i + 1, column=j + 1).value
            if value is None:
                value = ''
            exel_array[i].append(value)
    # return get_array(file_name=exel_otchet_file)
    return exel_array


def get_otchet_rows_dict():  # из названия файлов достает марку и делает приписку а так же достает новые коллизии
    # и неразрешенные коллизии а так же схлопывает строки если марки одинаковые
    added_marks = []
    rows_array = []
    data = get_exel_array()
    for i in data[1:len(data) - 1]:
        marks = i[0].split('.')
        i[0] = marks[len(marks) - 1]
        splited_mark = i[0].split('_')
        replaced_mark = f'{splited_mark[1]}_{splited_mark[0]}'

        if replaced_mark in added_marks:
            rows_array.append([f'Кол-во конфликтов между {replaced_mark}', i[1], i[4]])
            added_marks.append(replaced_mark)
        else:
            rows_array.append([f'Кол-во конфликтов между {i[0]}', i[1], i[4]])
            added_marks.append(i[0])
    for i in range(len(rows_array)):
        for j in range(3):
            if rows_array[i][j] == '' or rows_array[i][j] == 'Коллизий не обнаружено':
                rows_array[i][j] = 0
    marks_and_rows_dict = {}
    for value in rows_array:
        if value[0] not in marks_and_rows_dict.keys():
            marks_and_rows_dict.update({value[0]: [value[1], value[2]]})
        else:
            mark_value = marks_and_rows_dict[value[0]]
            marks_and_rows_dict[value[0]] = [int(mark_value[0]) + int(value[1]), int(mark_value[1]) + int(value[2])]
    return marks_and_rows_dict


def get_otchet_marks_array():  # достает только марки и делает приписку
    marks_array = []
    added_marks = []
    data = get_exel_array()
    for i in data[1:len(data) - 1]:
        marks = i[0].split('.')
        i[0] = marks[len(marks) - 1]
        splited_mark = i[0].split('_')
        replaced_mark = f'{splited_mark[1]}_{splited_mark[0]}'
        if replaced_mark in added_marks:
            marks_array.append(f'Кол-во конфликтов между {replaced_mark}')
            added_marks.append(replaced_mark)
        else:
            marks_array.append(f'Кол-во конфликтов между {i[0]}')
            added_marks.append(i[0])
    return list(set(marks_array))


def get_main_otchet_array():  # читает главный отчет и представляет его в виде двумерного массива
    wb = openpyxl.load_workbook(main_otchet_file, data_only=True)
    worksheet = wb['Лист1']
    exel_array = []
    for i in range(worksheet.max_row):
        exel_array.append([])
        for j in range(worksheet.max_column):
            value = worksheet.cell(row=i + 1, column=j + 1).value
            if value is None:
                value = ''
            exel_array[i].append(value)
    print(worksheet.max_row)
    print(worksheet.max_column)
    # return get_array(file_name=exel_otchet_file)
    return exel_array
#    return get_array(file_name=main_otchet_file)


def get_main_otchet_marks():  # возвращает марки из главного отчета
    return set(list(map(lambda x: x[0], get_main_otchet_array()))[2:])


def get_new_marks():  # возвращает новые марки
    marks = set(get_otchet_marks_array())
    main_marks = get_main_otchet_marks()
    difference = marks - main_marks
    return list(difference)


def get_rows_for_empty_list(worksheet):  # добавляет марки из отчета в главный отчет
    marks = get_otchet_marks_array()
    sorted_marks = sorted(marks)
    worksheet.cell(row=0 + 1, column=0 + 1).value = 'Дата'
    worksheet.cell(row=0 + 2, column=0 + 1).value = 'Конфликты'
    worksheet.cell(row=0 + 3, column=0 + 1).value = 'Итого:'
    for index, value in enumerate(sorted_marks):
        worksheet.cell(row=index + 4, column=0 + 1).value = value


def get_main_marks_and_rows():  # превращает главный exel отчет в пары ключ(марка): значение(строка после)
    main_rows_dict = {}
    for i in get_main_otchet_array():
        main_rows_dict.update({i[0]: i[1:]})
    return main_rows_dict


def moved_right_rows():  # добвляет в начале строки 2 пустых элемента в которые будут вписаны данные
    # новой проверки
    # wb = openpyxl.load_workbook(r"c:/users/skrut/onedrive/рабочий стол/exelexamples/kt101r_главный отчет.xlsx")
    # worksheet = wb['openpyxl']
    # wb.save(r"C:/Users/skrut/OneDrive/Рабочий стол/exelExamples/KT101R_Главный отчет.xlsx")
    moved_rows = get_main_marks_and_rows()

    for i in moved_rows.keys():
        moved_rows[i] = [0, 0] + moved_rows[i]
    new_marks = get_new_marks()
    if new_marks:
        for i in new_marks:
            moved_rows.update({i: [0, 0]})
    return moved_rows


def write_row(row, row_num, column_start, worksheet):  # функция написания строки в лист(worksheet) в строку номер (row
    # num) а сама
    # строка = row
    for i, value in enumerate(row):
        worksheet.cell(row=row_num, column=i + column_start).value = row[i]


def paint_row(row, row_num, color, worksheet, column_start):
    thin = Side(border_style="thin", color="2E3234")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i in range(len(row)):
        worksheet.cell(row=row_num, column=i + column_start).fill = PatternFill('solid',
                                                                                fgColor=color)
        worksheet.cell(row=row_num, column=i + column_start).border = border
    worksheet.cell(row=row_num, column=1).fill = PatternFill('solid',
                                                             fgColor=color)
    worksheet.cell(row=row_num, column=1).border = border


def write_if_main_is_empty():  # если главный отчет пустой - заполняет его марками в первом столбце(теми
    # которые были в начальных данных в алфавитном порядке марок) и вызывает функцию заполнения
    global main_otchet_file
    wb = openpyxl.load_workbook(main_otchet_file)
    worksheet = wb['Лист1']
    get_rows_for_empty_list(worksheet)
    wb.save(main_otchet_file)
    write_if_data_exists()


def write_if_data_exists():  # если главный отчет уже заполнен - добавляет данные по новой проверке
    # в начало списка проверок
    exel_array = get_exel_array()
    main_array = get_main_otchet_array()
    wb = openpyxl.load_workbook(main_otchet_file)
    worksheet = wb['Лист1']
    current_otchet_rows = get_otchet_rows_dict()
    result = exel_array.pop()
    line_1 = main_array[0]
    line_2 = main_array[1]
    line_3 = main_array[2]
    line_1.insert(1, '')
    line_1.insert(1, f'{datetime.date.today()}')
    line_2.insert(1, 'Конфликты')
    line_2.insert(1, 'Дата')
    line_3.insert(1, result[4])
    line_3.insert(1, result[1])
    moved_main_otchet_rows = moved_right_rows()
    sorted_otchet_marks = sorted(moved_main_otchet_rows.keys())
    keys_and_colors = {}
    counter = 0
    for index, value in enumerate(sorted_otchet_marks):
        if value in ['Дата', 'Конфликты', 'Итого:', '', None]:
            continue
        if value in current_otchet_rows.keys():
            moved_main_otchet_rows[value][0] = current_otchet_rows[value][0]
            moved_main_otchet_rows[value][1] = current_otchet_rows[value][1]
        else:
            moved_main_otchet_rows[value][0] = ''
            moved_main_otchet_rows[value][1] = ''
        if counter == len(array_of_colors):
            counter = 0
        worksheet.cell(row=index + 4, column=0 + 1).value = value
        keys_and_colors.update({value: array_of_colors[counter]})
        counter += 1

    for value in range(len(moved_main_otchet_rows.keys())):
        row_key = worksheet.cell(row=value + 4, column=0 + 1).value
        if row_key not in ['Дата', 'Конфликты', 'Итого:', '', None]:
            write_row(moved_main_otchet_rows[row_key], value + 4, 2, worksheet)
            paint_row(moved_main_otchet_rows[row_key], value + 4, keys_and_colors[row_key], worksheet, 2)
    write_row(line_1, 0 + 1, 1, worksheet)
    write_row(line_2, 1 + 1, 1, worksheet)
    write_row(line_3, 2 + 1, 1, worksheet)
    paint_row(line_1, 0 + 1, '92FF88', worksheet, 1)
    paint_row(line_2, 1 + 1, '92FF88', worksheet, 1)
    paint_row(line_3, 2 + 1, 'D9FF88', worksheet, 1)
    wb.save(main_otchet_file)


def write_data_in_main_otchet():  # заполняет данными в зависимости от того заполнен ли главный отчет
    # или он пуст
    main_array = get_main_otchet_array()
    for i in main_array:
        if not i:
            continue
        else:
            return write_if_data_exists
    return write_if_main_is_empty


write_data_in_main_otchet()()
'''
app = QApplication(sys.argv)
w = QWidget()
w.setWindowTitle('Добавление данных проверки')
w.resize(605, 185)


def get_main_dir(directory):
    global main_otchet_file
    global text_field_1
    main_otchet_file = directory
    text_field_1.setText(directory)


def get_otchet_dir(directory):
    global exel_otchet_file
    global text_field_2
    exel_otchet_file = directory
    text_field_2.setText(directory)


def start_recording(parent):
    global main_otchet_file
    global exel_otchet_file
    reply = QMessageBox.question(parent, 'Предупреждение',
                                 "Вы действительно хотите запустить генерацию отчета?\nПроверьте правильность "
                                 "указанного пути\nВнесенные изменения будут необратимы", QMessageBox.Yes |
                                 QMessageBox.No, QMessageBox.No)
    if reply == QMessageBox.Yes:
        main_otchet_file = get_xlsx_in_dir(main_otchet_file)[0]
        exel_otchet_file = get_xlsx_in_dir(exel_otchet_file)[0]
        print(main_otchet_file)
        print(exel_otchet_file)
        write_data_in_main_otchet()()
    else:
        pass


btn1 = QPushButton()
btn2 = QPushButton()
btn3 = QPushButton()

btn1.setText('Выбор папки с проектом')
btn1.setParent(w)
btn1.move(5, 5)
btn1.resize(175, 50)
btn1.clicked.connect(lambda: get_main_dir(QFileDialog.getExistingDirectory()))

btn3.setText('Выбор папки с проверкой')
btn3.setParent(w)
btn3.move(5, 60)
btn3.resize(175, 50)
btn3.clicked.connect(lambda: get_otchet_dir((QFileDialog.getExistingDirectory())))

btn2.setText('Запустить генерацию отчета')
btn2.setParent(w)
btn2.move(5, 115)
btn2.resize(595, 62)
btn2.clicked.connect(lambda: start_recording(w))

text_field_1 = QTextEdit()
text_field_1.setParent(w)
text_field_1.move(200, 15)
text_field_1.resize(400, 25)
text_field_1.show()

text_field_2 = QTextEdit()
text_field_2.setParent(w)
text_field_2.move(200, 70)
text_field_2.resize(400, 25)
text_field_2.show()

directory_button = QPushButton()
directory_button.setText("Выберите папку")
directory_button.setMinimumSize(QtCore.QSize(70, 40))

w.show()
sys.exit(app.exec())
'''